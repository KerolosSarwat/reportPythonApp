import sys
import os
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QDateEdit, QMessageBox, QTableView
from PyQt5.QtCore import Qt, QAbstractTableModel
import pandas as pd
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

class PatientData(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.patient_name = QLineEdit(self)
        self.phone_number = QLineEdit(self)
        self.operation_name = QLineEdit(self)
        self.cost = QLineEdit(self)
        self.paid = QLineEdit(self)
        self.date = QDateEdit(self)
        self.surgeon = QLineEdit(self)
        self.current_date = datetime.now().date()
        self.date.setDate(self.current_date)

        self.calculate_button = QPushButton('باقى الحساب', self)
        self.save_button = QPushButton('حفظ', self)
        self.display_button = QPushButton('عمليات غداً', self)

        # Create table view
        self.table_view = QTableView(self)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel('اسم المريض'))
        layout.addWidget(self.patient_name)
        layout.addWidget(QLabel('الموبايل'))
        layout.addWidget(self.phone_number)
        layout.addWidget(QLabel('العملية'))
        layout.addWidget(self.operation_name)
        layout.addWidget(QLabel('التكلفة'))
        layout.addWidget(self.cost)
        layout.addWidget(QLabel('المدفوع'))
        layout.addWidget(self.paid)
        layout.addWidget(QLabel('تاريخ العملية'))
        layout.addWidget(self.date)
        layout.addWidget(QLabel('الطبيب'))
        layout.addWidget(self.surgeon)
        layout.addWidget(self.calculate_button)
        layout.addWidget(self.save_button)
        layout.addWidget(self.display_button)
        layout.addWidget(self.table_view)

        self.calculate_button.clicked.connect(self.calculate_remaining_money)
        self.save_button.clicked.connect(self.save_data)
        self.display_button.clicked.connect(self.display_records)

        self.setWindowTitle('كشف حجوزات العمليات')
        self.setGeometry(300, 300, 900, 700)
        self.show()

    def calculate_remaining_money(self):
        cost = float(self.cost.text())
        paid = float(self.paid.text())
        remaining_money = cost - paid
        QMessageBox.information(self, 'باقى الحساب', f'باقى الحساب {remaining_money}')

    def save_data(self):
        patient_name = self.patient_name.text()
        phone_number = self.phone_number.text()
        operation_name = self.operation_name.text()
        cost = float(self.cost.text())
        paid = float(self.paid.text())
        date = self.date.date().toString('yyyy-MM-dd')
        surgeon = self.surgeon.text()
        remaining_money = cost - paid
        if not patient_name or not phone_number or not operation_name or not cost or not paid or not surgeon:
            QMessageBox.warning(self, "تحذير", "من فضلك تأكد من ملء جميع الحقول")
            return
        
        try:
            cost = float(self.cost.text())
            paid = float(self.paid.text())
            if cost <= 0:
                QMessageBox.warning(self, "خطأ", "يجب أن تكون قيمة التكلفة أكبر من الصفر")
                return
            if paid > cost:
                QMessageBox.warning(self, "خطأ", "قيمة المدفوع اكبر من التكلفة")
                return
        except ValueError:
            QMessageBox.warning(self, "خطأ", "يجب إدخال أرقام في حقلي التكلفة والسداد")
            return
    
        
        # Create a new DataFrame from input fields
        new_data = {
        "patient_name": [patient_name],
        "phone_number": [phone_number],
        "operation_name": [operation_name],
        "cost": [cost],
        "paid": [paid],
        "date": [date],
        "surgeon": [surgeon],
        "remaining_money": [remaining_money]
        }

        new_df = pd.DataFrame(new_data)

        # Load existing Excel file
        wb = load_workbook("daily_report.xlsx")
        ws = wb.active

        # Append new data to the last row of the Excel sheet
        last_row = ws.max_row + 1

        ws.cell(row=last_row, column=1, value=patient_name)
        ws.cell(row=last_row, column=2, value=phone_number)
        ws.cell(row=last_row, column=3, value=operation_name)
        ws.cell(row=last_row, column=4, value=cost)
        ws.cell(row=last_row, column=5, value=paid)
        ws.cell(row=last_row, column=6, value=date)
        ws.cell(row=last_row, column=7, value=surgeon)
        ws.cell(row=last_row, column=8, value=remaining_money)

        if not os.path.exists("daily_report.xlsx"):
            QMessageBox.warning(self, "خطأ", "ملف التقرير اليومي غير موجود")
            return
        # Save the updated Excel file
        wb.save("daily_report.xlsx")
        QMessageBox.information(self, "حفظ البيانات", "تم الحفظ بنجاح")

    def display_records(self):

        # Load existing Excel file
        df = pd.read_excel("daily_report.xlsx", sheet_name="Sheet")
        
        tomorrow = date.today() + pd.DateOffset(days=1)
        col_name = 'تاريخ العملية'
        df[col_name] = pd.to_datetime(df[col_name])

        newDF = df[df[col_name] == tomorrow.strftime("%Y-%m-%d")]

        # Create a table model from the DataFrame


        model = self.DataTableModel(newDF, df.columns.values)
        self.table_view.setModel(model)
    class DataTableModel(QAbstractTableModel):
        def __init__(self, data, hearders):
            super().__init__()
            self.data = data
            self.headers = hearders

        def rowCount(self, parent):
            return len(self.data)

        def columnCount(self, parent):
            return len(self.data.columns)

        def data(self, index, role=Qt.DisplayRole):
            if index.isValid():
                if role == Qt.DisplayRole:
                    value = self.data.iloc[index.row(), index.column()]
                    return str(value)

        def headerData(self, section, orientation: Qt.Orientation, role=Qt.DisplayRole):
            if role != Qt.DisplayRole:
                return None
            if orientation == Qt.Horizontal and role == Qt.DisplayRole:
                return self.headers[section]                
            return None
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = PatientData()
    sys.exit(app.exec_())